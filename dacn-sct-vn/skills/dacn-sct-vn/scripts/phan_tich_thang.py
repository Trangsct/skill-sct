#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
phan_tich_thang.py — Phân tích file Excel theo dõi sản xuất công nghiệp hằng tháng
của Phòng Quản lý Công nghiệp, Sở Công Thương tỉnh Lào Cai.

Mục đích: phục vụ báo cáo thực hiện Nghị quyết 169/NQ-CP (chỉ tiêu 3, 4, 15).

CẤU TRÚC FILE ĐẦU VÀO (mặc định, có thể chỉnh bằng tham số):
  - Sheet 'T1'..'T12': mỗi sheet là một tháng.
  - Dòng tiêu đề nằm ở khoảng dòng 2-4; dòng dữ liệu bắt đầu từ dòng 5.
  - Các cột (theo thứ tự trong file chuẩn):
        A: TT | B: Chỉ tiêu - Sản phẩm | C: ĐVT
        D: Kế hoạch năm | E: Sản xuất tháng trước
        F: Sản xuất trong tháng | G: Tiêu thụ trong tháng | H: Tồn kho cuối tháng
        I: So với sản xuất tháng trước | J: Luỹ kế đến tháng báo cáo
        K: So với kế hoạch năm | L: So cùng kỳ năm trước | M: Kế hoạch tháng sau

CÁCH DÙNG
  python3 phan_tich_thang.py <file.xlsx> --thang 5
  python3 phan_tich_thang.py <file.xlsx> --thang 5 --csv ketqua.csv
  python3 phan_tich_thang.py <file.xlsx> --thang 5 --chi-canh-bao

NGUYÊN TẮC BẤT BIẾN
  - Script CHỈ đọc số từ file, KHÔNG suy đoán, KHÔNG điền số thiếu.
  - Ô trống hoặc lỗi công thức (#DIV/0!, #REF!...) -> báo 'thiếu số liệu', không thay bằng 0.
  - Kết quả in ra là số liệu thô để người soạn văn bản kiểm tra lại, không phải số đã duyệt.
"""

import argparse
import csv
import sys
from datetime import date

try:
    import openpyxl
except ImportError:
    sys.exit("Thiếu thư viện openpyxl. Cài: pip install openpyxl --break-system-packages")


# ----- Cấu hình cột mặc định (0-based) -----
COT = {
    "tt": 0,
    "ten": 1,
    "dvt": 2,
    "ke_hoach_nam": 3,
    "sx_thang_truoc": 4,
    "sx_trong_thang": 5,
    "tieu_thu": 6,
    "ton_kho": 7,
    "so_thang_truoc": 8,
    "luy_ke": 9,
    "so_ke_hoach": 10,
    "so_cung_ky": 11,
}
DONG_BAT_DAU = 5

LOI_EXCEL = ("#DIV/0!", "#REF!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!")


def la_so(v):
    """Trả về float nếu ô là số dùng được, ngược lại None."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s in LOI_EXCEL:
            return None
        s = s.replace(" ", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def xep_hang_canh_bao(ty_le_hoan_thanh, thang):
    """
    Gắn nhãn cảnh báo theo reference 07 mục IV.
    ty_le_hoan_thanh: luỹ kế / kế hoạch năm, tính theo %.
    """
    if ty_le_hoan_thanh is None:
        return "THIEU-SO-LIEU", None
    tien_do_chuan = thang / 12.0 * 100.0
    do_lech = ty_le_hoan_thanh - tien_do_chuan
    if do_lech >= -3:
        muc = "AN-TOAN"
    elif do_lech >= -8:
        muc = "CAN-CHU-Y"
    elif do_lech >= -15:
        muc = "NGUY-CO"
    else:
        muc = "BAO-DONG"
    return muc, do_lech


def uoc_ca_nam(luy_ke, thang, theo_mua=False, luy_ke_cung_ky=None, ca_nam_truoc=None):
    """
    Ước thực hiện cả năm.
      - Mặc định: ngoại suy tuyến tính theo bình quân tháng (reference 07 mục III).
      - theo_mua=True: dùng tỷ lệ so cùng kỳ (áp dụng cho điện phát, nông sản chế biến).
    """
    if theo_mua:
        if luy_ke is None or luy_ke_cung_ky in (None, 0) or ca_nam_truoc is None:
            return None
        return ca_nam_truoc * (luy_ke / luy_ke_cung_ky)
    if luy_ke is None or thang == 0:
        return None
    return luy_ke / thang * 12.0


def doc_sheet(duong_dan, thang):
    wb = openpyxl.load_workbook(duong_dan, data_only=True)
    ten_sheet = f"T{thang}"
    if ten_sheet not in wb.sheetnames:
        sys.exit(
            f"Không tìm thấy sheet '{ten_sheet}'. Các sheet có trong file: {wb.sheetnames}"
        )
    return wb[ten_sheet]


def phan_tich(duong_dan, thang, dong_bat_dau=DONG_BAT_DAU):
    ws = doc_sheet(duong_dan, thang)
    ket_qua = []
    for r in ws.iter_rows(min_row=dong_bat_dau, max_row=ws.max_row, values_only=True):
        if r is None or len(r) <= COT["luy_ke"]:
            continue
        ten = r[COT["ten"]]
        if ten is None or str(ten).strip() == "":
            continue
        ten = str(ten).strip().replace("\n", " ")

        kh = la_so(r[COT["ke_hoach_nam"]])
        sx = la_so(r[COT["sx_trong_thang"]])
        lk = la_so(r[COT["luy_ke"]])
        so_ck = la_so(r[COT["so_cung_ky"]])
        so_tt = la_so(r[COT["so_thang_truoc"]])

        ty_le = (lk / kh * 100.0) if (lk is not None and kh not in (None, 0)) else None
        muc, do_lech = xep_hang_canh_bao(ty_le, thang)
        uoc = uoc_ca_nam(lk, thang)
        # Tỷ lệ so cùng kỳ trong file được ghi dạng hệ số (1,05 = tăng 5%)
        tang_cung_ky = (so_ck - 1.0) * 100.0 if so_ck is not None else None
        tang_thang_truoc = (so_tt - 1.0) * 100.0 if so_tt is not None else None

        ket_qua.append(
            {
                "tt": r[COT["tt"]],
                "ten": ten,
                "dvt": (str(r[COT["dvt"]]).strip() if r[COT["dvt"]] else ""),
                "ke_hoach_nam": kh,
                "sx_trong_thang": sx,
                "luy_ke": lk,
                "ty_le_kh": ty_le,
                "do_lech_tien_do": do_lech,
                "muc_canh_bao": muc,
                "uoc_ca_nam": uoc,
                "tang_so_cung_ky": tang_cung_ky,
                "tang_so_thang_truoc": tang_thang_truoc,
            }
        )
    return ket_qua


def dinh_dang(v, n=1):
    if v is None:
        return "—"
    return f"{v:,.{n}f}".replace(",", " ")


def in_bang(rows, chi_canh_bao=False, thang=None):
    uu_tien = {"BAO-DONG": 0, "NGUY-CO": 1, "CAN-CHU-Y": 2, "THIEU-SO-LIEU": 3, "AN-TOAN": 4}
    if chi_canh_bao:
        rows = [r for r in rows if r["muc_canh_bao"] in ("BAO-DONG", "NGUY-CO", "CAN-CHU-Y")]
    rows = sorted(rows, key=lambda r: (uu_tien.get(r["muc_canh_bao"], 9), r["ten"]))

    tien_do_chuan = thang / 12.0 * 100.0 if thang else None
    print()
    print("=" * 118)
    print(f"PHÂN TÍCH SẢN XUẤT CÔNG NGHIỆP — THÁNG {thang}/2026")
    if tien_do_chuan:
        print(f"Tiến độ chuẩn tại tháng {thang}: {tien_do_chuan:.1f}% kế hoạch năm")
    print("Nguồn: file theo dõi của Phòng QLCN. Số liệu thô — phải kiểm tra lại trước khi phát hành.")
    print("=" * 118)
    print(
        f"{'Sản phẩm':<38}{'ĐVT':<9}{'KH năm':>14}{'Luỹ kế':>14}"
        f"{'%KH':>8}{'Lệch':>8}{'±% CK':>9}  {'Mức'}"
    )
    print("-" * 118)
    for r in rows:
        print(
            f"{r['ten'][:37]:<38}{r['dvt'][:8]:<9}"
            f"{dinh_dang(r['ke_hoach_nam'], 0):>14}"
            f"{dinh_dang(r['luy_ke'], 0):>14}"
            f"{dinh_dang(r['ty_le_kh']):>8}"
            f"{dinh_dang(r['do_lech_tien_do']):>8}"
            f"{dinh_dang(r['tang_so_cung_ky']):>9}  "
            f"{r['muc_canh_bao']}"
        )
    print("-" * 118)

    dem = {}
    for r in rows:
        dem[r["muc_canh_bao"]] = dem.get(r["muc_canh_bao"], 0) + 1
    print("TỔNG HỢP: " + " | ".join(f"{k}: {v}" for k, v in sorted(dem.items())))
    print()
    print("GHI CHÚ NGHIỆP VỤ:")
    print("  - 'Lệch' = %KH thực tế trừ tiến độ chuẩn (điểm %). Âm là chậm.")
    print("  - '±% CK' = tăng/giảm so cùng kỳ năm trước, suy từ hệ số trong file.")
    print("  - Dòng 'Điện phát' và nhóm nông sản chế biến CHỊU TÍNH MÙA VỤ:")
    print("    không kết luận theo cột %KH, phải ngoại suy theo cùng kỳ (reference 07 mục III).")
    print("  - 'Điện phát' KHÁC 'Điện thương phẩm' (chỉ tiêu 15) — xem reference 04 mục III.")
    print("  - Mọi dòng THIEU-SO-LIEU phải bổ sung trước khi đưa vào báo cáo.")
    print()


def xuat_csv(rows, duong_dan):
    truong = [
        "tt", "ten", "dvt", "ke_hoach_nam", "sx_trong_thang", "luy_ke",
        "ty_le_kh", "do_lech_tien_do", "muc_canh_bao", "uoc_ca_nam",
        "tang_so_cung_ky", "tang_so_thang_truoc",
    ]
    with open(duong_dan, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=truong)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k) for k in truong})
    print(f"Đã ghi CSV: {duong_dan}")


def main():
    p = argparse.ArgumentParser(
        description="Phân tích số liệu sản xuất công nghiệp hằng tháng (NQ 169)."
    )
    p.add_argument("file", help="Đường dẫn file Excel theo dõi")
    p.add_argument("--thang", type=int, required=True, help="Tháng báo cáo (1-12)")
    p.add_argument("--dong-bat-dau", type=int, default=DONG_BAT_DAU,
                   help=f"Dòng dữ liệu đầu tiên (mặc định {DONG_BAT_DAU})")
    p.add_argument("--csv", help="Xuất kết quả ra file CSV")
    p.add_argument("--chi-canh-bao", action="store_true",
                   help="Chỉ hiển thị sản phẩm ở mức CAN-CHU-Y trở lên")
    a = p.parse_args()

    if not 1 <= a.thang <= 12:
        sys.exit("Tháng phải từ 1 đến 12.")

    rows = phan_tich(a.file, a.thang, a.dong_bat_dau)
    if not rows:
        sys.exit("Không đọc được dòng dữ liệu nào. Kiểm tra lại --dong-bat-dau và cấu trúc file.")
    in_bang(rows, a.chi_canh_bao, a.thang)
    if a.csv:
        xuat_csv(rows, a.csv)


if __name__ == "__main__":
    main()
